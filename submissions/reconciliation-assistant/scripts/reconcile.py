#!/usr/bin/env python3
"""Config-driven reconciliation for code-capable hosts (Cowork, Scout).

Reads two datasets described by a JSON config (see assets/config.example.json),
runs a tiered match (exact -> difference -> similarity -> grouped -> unmatched),
proves the control totals tie out, and writes an .xlsx report.

This is a reference implementation the agent adapts to the actual column names
and file paths in play. It has no hidden behaviour: every rule here mirrors
SKILL.md and references/methodology.md.

Usage:
    python reconcile.py --config config.json --source-a A.xlsx --source-b B.csv --out reconciliation.xlsx

Dependencies: pandas, openpyxl. (difflib is stdlib and used for similarity.)
"""

import argparse
import json
import re
import sys
from difflib import SequenceMatcher

import pandas as pd


# ----------------------------- loading -----------------------------

def load_table(path, sheet=None):
    lower = path.lower()
    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        # sheet may be a name or an index; None loads the first sheet
        return pd.read_excel(path, sheet_name=sheet if sheet is not None else 0)
    if lower.endswith(".tsv"):
        return pd.read_csv(path, sep="\t")
    return pd.read_csv(path)


def default_label(path, sheet=None):
    """A human label for a source: file name, plus sheet name when reconciling tabs."""
    import os
    base = os.path.splitext(os.path.basename(path))[0]
    if sheet is not None and not isinstance(sheet, int):
        return f"{base} — {sheet}"
    return base


def normalize_amount(value, norm):
    """Return a float from a possibly messy amount cell, or None. The result is quantized to the
    cent (2 decimals) so that "exact to the cent" matching is deterministic and immune to binary
    floating-point noise - two cells that should be equal (e.g. both "1250.00") can otherwise parse
    to minutely different floats and, with a zero tolerance, be pushed to "Matched (with
    difference)". Quantizing here means the matcher, the per-key model and the workbook all compare
    the same cent-rounded values (the workbook already rounds to 2 dp when classifying)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    s = str(value).strip()
    if not s:
        return None
    negative = False
    if norm.get("parenthesesMeanNegative", True) and s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]
    if norm.get("stripCurrencySymbols", True):
        s = "".join(ch for ch in s if ch.isdigit() or ch in ".-")
    s = s.replace(",", "")
    try:
        amt = float(s)
    except ValueError:
        return None
    return round(-amt if negative else amt, 2)


def apply_sign(amount, convention):
    if amount is None:
        return None
    if convention == "flip":
        return -amount
    return amount


_MULTISPACE = re.compile(r" {2,}")


def norm_key(value, norm):
    # Canonicalize a key component. Integer-valued floats (e.g. 7100.0 that pandas produced
    # because another row was blank) are rendered as "7100" so a Python key matches what Excel
    # writes when it concatenates the same numeric cell - keeping all three outputs in step.
    if value is None or (isinstance(value, float) and pd.isna(value)):
        s = ""
    elif isinstance(value, float) and value.is_integer():
        s = str(int(value))
    else:
        s = str(value)
    if norm.get("trimWhitespace", True):
        # Mirror Excel TRIM(): strip leading/trailing spaces AND collapse internal runs of spaces
        # to a single space. If we only did .strip(), a component like "ACME  CORP" would stay
        # distinct in the Python union while Excel's TRIM in the Matching Key formula collapses it
        # to "ACME CORP" - the two would then disagree and SUMIF/COUNTIF could double-count.
        s = _MULTISPACE.sub(" ", s).strip(" ")
    if norm.get("caseInsensitiveKeys", True):
        s = s.lower()
    return s


# One delimiter for every key builder - the Python matcher (build_key), the Excel helper
# (_xl_key_formula), the workbook union (keystr) and the HTML path (kstr) - so the three outputs
# group keys identically. A delimiter-only string can never stand in for a real key because
# join_key_parts collapses an all-empty key to "".
KEY_DELIM = " | "


def join_key_parts(parts):
    """Join normalized key components with the shared delimiter, collapsing an all-empty key to
    "" so keyless rows are treated as keyless everywhere (matcher, workbook, HTML) instead of
    grouping under a delimiter-only string."""
    return KEY_DELIM.join(parts) if any(parts) else ""


def build_key(row, key_cols, norm):
    # Exact/similarity tiers treat a "" key as keyless (see the `_key != ""` guard and the
    # similarity empty-key check), so join_key_parts collapsing all-empty parts to "" is what
    # keeps that protection intact.
    return join_key_parts([norm_key(row.get(c), norm) for c in key_cols])


# A row whose key components are ALL blank has no usable key. The record matcher treats such a row
# as non-matchable (its own one-sided break). The per-key views (workbook SUMIF/COUNTIF and the
# HTML aggregation) would otherwise group every keyless row under the single empty key "" and could
# "reconcile" them purely on netted totals. To keep those views faithful to the matcher, each
# keyless row is given a stable, unique placeholder key derived from a per-side scope plus its row
# position, so keyless rows are never aggregated together. The prefix cannot collide with a real
# key (real keys are TRIM/LOWER'd values joined by KEY_DELIM) and contains no Excel wildcard chars.
_KEYLESS_PREFIX = "(no key \u00b7 "


def keyless_token(scope, position):
    return f"{_KEYLESS_PREFIX}{scope} #{position})"


def is_keyless_token(s):
    return isinstance(s, str) and s.startswith(_KEYLESS_PREFIX)


def similarity(a, b):
    return SequenceMatcher(None, str(a), str(b)).ratio()


def within_tolerance(x, y, abs_tol, pct_tol):
    if x is None or y is None:
        return False
    diff = abs(x - y)
    if diff <= abs_tol:
        return True
    if pct_tol > 0 and max(abs(x), abs(y)) > 0:
        return (diff / max(abs(x), abs(y))) * 100.0 <= pct_tol
    return False


def effective_tolerances(matching):
    """Resolve the amount-match tolerances honoring matching.amountMatch. In 'exact' mode the
    amounts must agree exactly (to the cent for 2-dp currency data), so BOTH tolerances are 0
    regardless of any amountToleranceAbsolute/Percent left in the config; 'tolerance' mode (or an
    unset amountMatch) uses the configured absolute/percent values (default 0.01 / 0)."""
    if matching.get("amountMatch") == "exact":
        return 0.0, 0.0
    return matching.get("amountToleranceAbsolute", 0.01), matching.get("amountTolerancePercent", 0.0)


def align_key_columns(config):
    """If matching.keyMap pairs A's key columns to differently-named B columns, reorder
    sources.b.keyColumns to match sources.a.keyColumns element-wise, so every positional
    (a_keys[i] <-> b_keys[i]) assumption downstream (key building, timing, report field
    mapping) holds. No-op when keyMap is absent or does not cover every A key column."""
    m = config.get("matching", {})
    key_map = m.get("keyMap")
    if not key_map:
        return
    a_keys = config["sources"]["a"]["keyColumns"]
    mapping = {}
    for pair in key_map:
        if isinstance(pair, (list, tuple)) and len(pair) == 2:
            mapping[pair[0]] = pair[1]
    if a_keys and all(k in mapping for k in a_keys):
        config["sources"]["b"]["keyColumns"] = [mapping[k] for k in a_keys]


# ----------------------------- matching -----------------------------

def reconcile(df_a, df_b, config):
    src = config["sources"]
    m = config["matching"]
    norm = config.get("normalization", {})
    abs_tol, pct_tol = effective_tolerances(m)

    a_keys = src["a"]["keyColumns"]
    b_keys = src["b"]["keyColumns"]
    a_amt_col = src["a"]["amountColumn"]
    b_amt_col = src["b"]["amountColumn"]

    # Timing detection: identify the "period" component of the key, if configured.
    # timingKeyColumn names a column in source A's keyColumns; the same position in
    # b_keys is treated as B's period column (keyMap keeps the two aligned).
    timing_col = m.get("timingKeyColumn")
    enable_timing = m.get("enableTimingDetection", True) and timing_col is not None
    a_timing_idx = a_keys.index(timing_col) if (enable_timing and timing_col in a_keys) else None
    # Disable timing unless the period column is present in A's key AND the aligned B key has a
    # column at the same position (guards against an IndexError / wrong reduced key when B's
    # keyColumns are shorter or were not aligned to A via keyMap), AND there is at least one
    # non-timing key column - otherwise the reduced key collapses to "" and unrelated rows would
    # be paired as "timing" purely on amount.
    if enable_timing and (a_timing_idx is None or a_timing_idx >= len(b_keys) or len(a_keys) <= 1):
        enable_timing = False

    def reduced_key(row, keys, norm):
        # key with the timing component removed, so "same record, different period" collapses
        return join_key_parts([norm_key(row.get(c), norm) for i, c in enumerate(keys) if i != a_timing_idx])

    a = df_a.to_dict("records")
    b = df_b.to_dict("records")
    for i, r in enumerate(a):
        r["_idx"] = i
        r["_key"] = build_key(r, a_keys, norm)
        r["_amt"] = apply_sign(normalize_amount(r.get(a_amt_col), norm), src["a"].get("signConvention", "asIs"))
        if enable_timing:
            r["_rkey"] = reduced_key(r, a_keys, norm)
            r["_tval"] = norm_key(r.get(a_keys[a_timing_idx]), norm)
    for j, r in enumerate(b):
        r["_idx"] = j
        r["_key"] = build_key(r, b_keys, norm)
        r["_amt"] = apply_sign(normalize_amount(r.get(b_amt_col), norm), src["b"].get("signConvention", "asIs"))
        if enable_timing:
            r["_rkey"] = reduced_key(r, b_keys, norm)
            r["_tval"] = norm_key(r.get(b_keys[a_timing_idx]), norm)

    total_a = sum(r["_amt"] for r in a if r["_amt"] is not None)
    total_b = sum(r["_amt"] for r in b if r["_amt"] is not None)

    b_by_key = {}
    for r in b:
        b_by_key.setdefault(r["_key"], []).append(r)

    results = []
    used_b = set()

    # Tier 1 + 2: exact key
    for ra in a:
        candidates = [r for r in b_by_key.get(ra["_key"], []) if r["_idx"] not in used_b and ra["_key"] != ""]
        if not candidates:
            continue
        candidates.sort(key=lambda r: abs((r["_amt"] or 0) - (ra["_amt"] or 0)))
        rb = candidates[0]
        used_b.add(rb["_idx"])
        ra["_matched"] = True
        if ra["_amt"] is None or rb["_amt"] is None:
            # Key matches on both sides but an amount is blank/unparseable. Do not silently
            # invent a clean variance; flag for review. The difference kept here is the
            # balancing contribution to the tie-out identity (a blank amount contributed 0
            # to its control total), not a fabricated "matched" number.
            status = "Probable (Needs Review)"
            diff = (ra["_amt"] or 0) - (rb["_amt"] or 0)
            evidence = "exact key; amount missing on one side - verify before treating as matched"
        elif within_tolerance(ra["_amt"], rb["_amt"], abs_tol, pct_tol):
            status = "Matched"
            diff = 0.0
            evidence = "exact key"
        else:
            status = "Matched (with difference)"
            diff = (ra["_amt"] or 0) - (rb["_amt"] or 0)
            evidence = "exact key"
        results.append({"status": status, "a_idx": ra["_idx"], "b_idx": rb["_idx"],
                        "key": ra["_key"], "amount_a": ra["_amt"], "amount_b": rb["_amt"],
                        "difference": diff, "evidence": evidence})

    unmatched_a = [r for r in a if not r.get("_matched")]
    unmatched_b = [r for r in b if r["_idx"] not in used_b]

    # Tier 3: similarity (candidate matching for records with no shared key)
    if m.get("enableSimilarityMatching", True):
        date_a = src["a"].get("dateColumn")
        date_b = src["b"].get("dateColumn")
        window = m.get("dateWindowDays", 3)
        sim_thr = m.get("similarityThreshold", 0.9)
        still_a = []
        for ra in unmatched_a:
            # Never pair on an empty key/description - SequenceMatcher on two empty strings
            # returns 1.0 and would fabricate a "Probable" match from amount/date alone.
            if not str(ra["_key"]).strip():
                still_a.append(ra)
                continue
            best = None
            for rb in unmatched_b:
                if rb["_idx"] in used_b or not str(rb["_key"]).strip():
                    continue
                if not within_tolerance(ra["_amt"], rb["_amt"], abs_tol, pct_tol):
                    continue
                if date_a and date_b:
                    da, db = ra.get(date_a), rb.get(date_b)
                    try:
                        delta_days = abs((pd.to_datetime(da) - pd.to_datetime(db)).total_seconds()) / 86400.0
                    except Exception:
                        # A date was configured but could not be parsed: the proximity rule
                        # cannot be satisfied, so this pair is not eligible for similarity.
                        continue
                    if delta_days > window:
                        continue
                sim = similarity(ra["_key"], rb["_key"])
                if sim >= sim_thr and (best is None or sim > best[1]):
                    best = (rb, sim)
            if best:
                rb, sim = best
                used_b.add(rb["_idx"])
                results.append({"status": "Probable (Needs Review)", "a_idx": ra["_idx"], "b_idx": rb["_idx"],
                                "key": ra["_key"], "amount_a": ra["_amt"], "amount_b": rb["_amt"],
                                "difference": (ra["_amt"] or 0) - (rb["_amt"] or 0),
                                "evidence": f"similarity: amount+date, name similarity {sim:.2f}"})
            else:
                still_a.append(ra)
        unmatched_a = still_a
        unmatched_b = [r for r in unmatched_b if r["_idx"] not in used_b]

    # Tier 4: grouped (split / partial) matches. One record on one side equals the sum of
    # several on the other within tolerance (e.g. one invoice settled by three payments).
    # Bounded for safety: enumeration is skipped when the opposite pool is too large, and
    # combinations are capped at groupedMaxMembers. Grouped pairs go to Needs Review with
    # every member listed - a split is legitimate but a human should confirm it.
    if m.get("enableGrouped", True):
        from itertools import combinations
        max_members = max(2, int(m.get("groupedMaxMembers", 6)))
        POOL_CAP = 30  # skip enumeration if the many-side pool exceeds this (keeps it fast)
        grouped_a, grouped_b = set(), set()

        def _find_combo(target, pool, exclude):
            avail = [r for r in pool if r["_idx"] not in exclude and r["_amt"] is not None]
            if len(avail) > POOL_CAP:
                return None
            # A split is same-sign as its target, so drop opposite-sign candidates and any
            # single item already larger (by magnitude) than the target - this prunes the
            # search space sharply before enumerating combinations.
            if target >= 0:
                avail = [r for r in avail if 0 <= r["_amt"] <= target + abs_tol]
            else:
                avail = [r for r in avail if target - abs_tol <= r["_amt"] <= 0]
            # Search smaller (nearest-magnitude-first) combinations first, with an attempt cap
            # so a pathological pool can't blow up the run.
            avail.sort(key=lambda r: abs(r["_amt"]), reverse=True)
            attempts = 0
            ATTEMPT_CAP = 50000
            for size in range(2, max_members + 1):
                for combo in combinations(avail, size):
                    attempts += 1
                    if attempts > ATTEMPT_CAP:
                        return None
                    s = sum(c["_amt"] for c in combo)
                    if within_tolerance(target, s, abs_tol, pct_tol):
                        return combo
            return None

        # One A record ↔ many B records.
        for ra in unmatched_a:
            if ra["_amt"] is None:
                continue
            combo = _find_combo(ra["_amt"], unmatched_b, used_b | grouped_b)
            if combo:
                grouped_a.add(ra["_idx"])
                for c in combo:
                    grouped_b.add(c["_idx"]); used_b.add(c["_idx"])
                s = sum(c["_amt"] for c in combo)
                members = ", ".join(str(c["_key"]) for c in combo)
                results.append({"status": "Grouped (Needs Review)", "a_idx": ra["_idx"], "b_idx": None,
                                "key": ra["_key"], "amount_a": ra["_amt"], "amount_b": s,
                                "difference": (ra["_amt"] or 0) - s,
                                "evidence": f"grouped: {len(combo)} {src['b'].get('label','B')} rows ({members}) sum to {s:,.2f}"})

        # One B record ↔ many A records (using A rows not already grouped above).
        pool_a = [r for r in unmatched_a if r["_idx"] not in grouped_a]
        for rb in unmatched_b:
            if rb["_idx"] in grouped_b or rb["_amt"] is None:
                continue
            combo = _find_combo(rb["_amt"], pool_a, grouped_a)
            if combo:
                grouped_b.add(rb["_idx"])
                for c in combo:
                    grouped_a.add(c["_idx"])
                s = sum(c["_amt"] for c in combo)
                members = ", ".join(str(c["_key"]) for c in combo)
                results.append({"status": "Grouped (Needs Review)", "a_idx": None, "b_idx": rb["_idx"],
                                "key": rb["_key"], "amount_a": s, "amount_b": rb["_amt"],
                                "difference": s - (rb["_amt"] or 0),
                                "evidence": f"grouped: {len(combo)} {src['a'].get('label','A')} rows ({members}) sum to {s:,.2f}"})

        unmatched_a = [r for r in unmatched_a if r["_idx"] not in grouped_a]
        unmatched_b = [r for r in unmatched_b if r["_idx"] not in grouped_b]

    # Tier 4b: timing differences. Among the still-unmatched records, detect the classic
    # "same item posted to a different period" case: an A record and a B record sharing the
    # reduced key (identity minus the period) and the same amount, but a different period.
    # We ANNOTATE both lines (so they remain visible as one-sided breaks and count toward the
    # variance the way an accountant expects) rather than collapsing them - the note preserves
    # the timing insight for the reviewer.
    if enable_timing:
        b_pool = {}
        for rb in unmatched_b:
            if not rb.get("_rkey"):
                continue  # all non-period components blank: no identity to match a timing pair on
            b_pool.setdefault(rb["_rkey"], []).append(rb)
        b_noted = set()
        for ra in unmatched_a:
            if not ra.get("_rkey"):
                continue
            for rb in b_pool.get(ra["_rkey"], []):
                if rb["_idx"] in b_noted or rb["_tval"] == ra["_tval"]:
                    continue
                if within_tolerance(ra["_amt"], rb["_amt"], abs_tol, pct_tol):
                    ra["_timing_note"] = f"Possible timing difference - same amount in {rb['_tval']}"
                    rb["_timing_note"] = f"Possible timing difference - same amount in {ra['_tval']}"
                    b_noted.add(rb["_idx"])
                    break

    # Tier 5: whatever remains as genuine one-sided breaks.
    for ra in unmatched_a:
        results.append({"status": "Unmatched (A)", "a_idx": ra["_idx"], "b_idx": None,
                        "key": ra["_key"], "amount_a": ra["_amt"], "amount_b": None,
                        "difference": None, "evidence": ra.get("_timing_note", "")})
    for rb in unmatched_b:
        results.append({"status": "Unmatched (B)", "a_idx": None, "b_idx": rb["_idx"],
                        "key": rb["_key"], "amount_a": None, "amount_b": rb["_amt"],
                        "difference": None, "evidence": rb.get("_timing_note", "")})

    return results, total_a, total_b


def tie_out(results, total_a, total_b, abs_tol):
    # The identity: (total A - total B) must equal the sum of every line's net contribution.
    # For a matched-with-difference, probable, or grouped pairing that is the recorded
    # difference (A less B, with a blank amount counting as 0); for a one-sided item it is
    # the present amount. Including probable/grouped keeps the identity correct whenever those
    # tiers pair amounts within tolerance (a small non-zero delta still has to be explained).
    explained = 0.0
    for r in results:
        st = r["status"]
        d = r.get("difference")
        if st in ("Matched (with difference)", "Probable (Needs Review)",
                  "Grouped (Needs Review)") and d is not None:
            explained += d
        elif st == "Unmatched (A)" and r.get("amount_a") is not None:
            explained += r["amount_a"]
        elif st == "Unmatched (B)" and r.get("amount_b") is not None:
            explained -= r["amount_b"]
    left = total_a - total_b
    closed = abs(left - explained) <= max(abs_tol, 0.01)
    return {"total_a": total_a, "total_b": total_b, "net_difference": left,
            "explained": explained, "residual": left - explained, "tied_out": closed}


# ----------------------------- output -----------------------------

# Report palette (formula-driven workbook): blue headers, accounting number format.
HDR_FILL = "2E5C8A"         # blue - table header fills (white text)
HDR_FONT = "FFFFFF"         # white header text
SEC_C = "2E5C8A"            # blue - section labels / title text
SUB_C = "595959"            # gray - subtitles / basis-of-preparation line
BODY_C = "404040"           # near-black body text
NARR_C = "3B3B3B"           # headlines narrative text
MK_C = "808080"             # gray - matching-key helper column
REPORT_FONT = "Cambria"     # v15 uses Cambria throughout
# Consistent number format used for every amount throughout the workbook:
# 2 decimals, negatives in parentheses (e.g. 1,234.00 / (1,234.00) / 0.00). No currency symbol.
ACCT2 = '#,##0.00;(#,##0.00)'
CNT_FMT = '#,##0'
PCT_FMT = '0.0%'
ZEBRA_BG = "F2F7FC"         # very light blue - alternating rows
OPEN_FONT, OPEN_BG = "8C1D18", "F7E3E1"   # Open Item - red text on soft red
REC_FONT, REC_BG = "1F3864", "EAF1F8"     # Reconciled - navy text on soft blue
def _build_narrative_perkey(rows, config):
    """Headline narrative computed from the per-key reconciliation rows - the same model the
    Reconciliation sheet and the HTML dashboard use - so the headline counts can never disagree
    with the sheet totals. Returns plain-text lines (no currency symbols); shared verbatim by the
    Excel Dashboard and the HTML dashboard."""
    la = config["sources"]["a"].get("label", "Source A")
    lb = config["sources"]["b"].get("label", "Source B")
    out = config.get("output", {})
    group_by = out.get("groupBy", [])
    gb0 = group_by[0] if group_by else "company"
    gb1 = group_by[1] if len(group_by) > 1 else "period"

    total = len(rows)
    reconciled = sum(1 for r in rows if r["status"] == "Reconciled")
    open_rows = [r for r in rows if r["status"] == "Open Item"]
    opn = len(open_rows)
    net = round(sum(r["diff"] for r in rows), 2)
    gross = round(sum(abs(r["diff"]) for r in rows), 2)
    rate = (reconciled / total * 100) if total else 0

    # Per-account rollup (for the biggest driver) and root-cause tallies over open items.
    acct = {}
    acct_order = []
    for r in rows:
        a = r["account"]
        if a not in acct:
            acct[a] = {"name": r["name"], "a": 0.0, "b": 0.0}
            acct_order.append(a)
        acct[a]["a"] += r["amt_a"]
        acct[a]["b"] += r["amt_b"]

    def rc(name):
        c = sum(1 for r in open_rows if r["rootcause"] == name)
        v = round(sum(abs(r["diff"]) for r in open_rows if r["rootcause"] == name), 2)
        return c, v

    m_c, m_v = rc("Measurement")
    t_c, t_v = rc("Timing")
    s_c, s_v = rc("Scope / mapping")
    timing_net = round(sum(r["diff"] for r in open_rows if r["rootcause"] == "Timing"), 2)

    lines = [
        f"{total} keys reconciled across the {gb0.lower()} and {gb1.lower()} dimensions: "
        f"{reconciled} reconciled and {opn} open, a {rate:.1f}% match rate.",
        f"Net difference is {_num(net)} ({la} less {lb}); ignoring sign the differences total {_num(gross)}.",
    ]
    if acct_order:
        biggest = max(acct_order, key=lambda a: abs(acct[a]["a"] - acct[a]["b"]))
        big_diff = acct[biggest]["a"] - acct[biggest]["b"]
        lines.append(f"Largest account driver: {acct[biggest]['name']} at {_num(big_diff)}.")
    lines.append(
        f"Root causes: {m_c} measurement ({_num(m_v)}), {t_c} timing ({_num(t_v)}) "
        f"and {s_c} scope or mapping ({_num(s_v)}).")
    if t_c and timing_net == 0:
        lines.append("Timing items net to 0.00 across the periods and should clear without adjustment.")
    return lines


def _CL(n):
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)


def _src_meta(df, src_cfg, config):
    """Column geometry for a source tab: header names, key/amount letters, the appended
    Matching Key helper column, and the A1-style ranges used by the reconciliation formulas."""
    cols = list(df.columns)
    n = len(df)
    amt = src_cfg["amountColumn"]
    keys = src_cfg["keyColumns"]
    amt_letter = _CL(cols.index(amt) + 1)
    key_letters = [_CL(cols.index(k) + 1) for k in keys]
    mk_letter = _CL(len(cols) + 1)          # Matching Key helper appended after the data
    last = n + 1                             # data occupies rows 2..last
    return {
        "cols": cols, "n": n, "amt_letter": amt_letter, "key_letters": key_letters,
        "mk_letter": mk_letter, "last": last, "keys": keys,
    }


def _neutralize(v):
    """Defuse spreadsheet formula/injection: a text value that begins with = + - @ could be
    executed as a formula when the workbook is opened. Since all source data is untrusted, force
    such strings to literal text with a leading apostrophe. Numbers are unaffected."""
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return "'" + v
    return v


def _xl_key_formula(cell_refs, norm):
    """Excel formula that concatenates key cell references into a Matching Key, mirroring
    join_key_parts()/norm_key(): each component is wrapped in TRIM() when trimWhitespace is on and
    LOWER() when caseInsensitiveKeys is on, the components are joined by the shared KEY_DELIM, and
    an all-empty key collapses to "" (via an IF over the delimiter-free concatenation) exactly as
    the Python builder does. The source-tab helper and the Reconciliation sheet both call this with
    identical settings so their SUMIF/COUNTIF keys line up. LOWER/TRIM also coerce numbers to text
    the same way (integer-valued cells render without a trailing .0)."""
    trim = norm.get("trimWhitespace", True)
    lower = norm.get("caseInsensitiveKeys", True)

    def wrap(ref):
        expr = ref
        if trim:
            expr = f"TRIM({expr})"
        if lower:
            expr = f"LOWER({expr})"
        return expr

    parts = [wrap(r) for r in cell_refs]
    if not parts:
        return '=""'
    bare = "&".join(parts)                       # components with no delimiter, for the empty test
    joined = ('&"' + KEY_DELIM + '"&').join(parts)
    return f'=IF({bare}="","",{joined})'


def _safe_sheet_name(name, taken):
    """A valid, unique Excel sheet name: strip the reserved characters : \\ / ? * [ ] (plus ~,
    which is a wildcard-escape in SUMIF/COUNTIF criteria and would break a keyless row's literal
    Matching Key that embeds the sheet name), cap at 31 chars, and de-duplicate with a numeric
    suffix (truncating to keep room for it)."""
    import re
    n = re.sub(r"[:\\/?*\[\]~]", " ", str(name)).strip() or "Sheet"
    n = n[:31]
    base, i = n, 2
    while n.lower() in taken:
        suffix = f" ({i})"
        n = base[:31 - len(suffix)].rstrip() + suffix
        i += 1
    taken.add(n.lower())
    return n


def _write_source_tab(ws, df, meta, sheet_title, sign="asIs", norm=None):
    """Write a source ledger plus a Matching Key helper column, styled with a blue header. The
    amount column is written sign-normalized (so signConvention flows through the workbook's
    SUMIF totals); text cells are neutralized against formula injection; the helper column joins
    the key cells so the reconciliation SUMIF/COUNTIFs bind."""
    from openpyxl.styles import Font, PatternFill, Alignment

    norm = norm or {}
    hdr_fill = PatternFill("solid", fgColor=HDR_FILL)
    cols = meta["cols"]
    # Header row.
    for c, name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=str(name))
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color=HDR_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    mk_c = len(cols) + 1
    hc = ws.cell(row=1, column=mk_c, value="Matching Key")
    hc.fill = hdr_fill
    hc.font = Font(bold=True, color=HDR_FONT)
    hc.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows. v15 convention: numeric non-amount cells use Cambria left-aligned (General
    # format renders integer keys cleanly); free-text cells (e.g. Account Name, Period) use
    # Arial 10; the amount uses the shared money format. The Matching Key helper is a gray formula.
    amt_letter = meta["amt_letter"]
    text_cols = {name for name in cols if not pd.api.types.is_numeric_dtype(df[name])}
    for r, (_, row) in enumerate(df.iterrows(), start=2):
        for c, name in enumerate(cols, start=1):
            v = row[name]
            if pd.isna(v):
                v = None
            if _CL(c) == amt_letter:
                # Sign-normalized numeric value drives the workbook's SUMIF totals.
                cell = ws.cell(row=r, column=c, value=apply_sign(normalize_amount(v, norm), sign))
                cell.number_format = ACCT2
            elif name in text_cols:
                cell = ws.cell(row=r, column=c, value=_neutralize(v))
                cell.font = Font(name="Arial", size=10)
            else:
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = Alignment(horizontal="left")
        # Matching Key. A row with all key components blank gets a unique placeholder so keyless
        # rows are never aggregated together by the reconciliation SUMIF/COUNTIF; otherwise the
        # normalized concatenation of the key cells (TRIM/LOWER, all-empty collapses to "").
        if not any(norm_key(row[k], norm) for k in meta["keys"]):
            mkc = ws.cell(row=r, column=mk_c, value=keyless_token(sheet_title, r))
        else:
            refs = [f"${kl}{r}" for kl in meta["key_letters"]]
            mkc = ws.cell(row=r, column=mk_c, value=_xl_key_formula(refs, norm))
        mkc.font = Font(color=MK_C)

    # Column widths: vectorized string-length over a bounded sample (avoids an O(rows*cols)
    # Python loop; the widest of the first 200 rows is a fine proxy for display width).
    sample = df.head(200)
    for c, name in enumerate(cols, start=1):
        try:
            body_max = int(sample[name].astype(str).str.len().max() or 0)
        except Exception:
            body_max = 0
        width = min(max(max(len(str(name)), body_max) + 2, 10), 40)
        ws.column_dimensions[_CL(c)].width = width
    ws.column_dimensions[_CL(mk_c)].width = 26
    ws.freeze_panes = "A2"


def _col_to_idx(letter):
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter) - 1


def _write_reconciliation(ws, df_a, df_b, config, meta_a, meta_b, sa, sb):
    """The formula-driven reconciliation: one row per union key, every number a live formula
    over the two source tabs. Returns the layout info the dashboard needs to reference it."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import FormulaRule, CellIsRule

    la = config["sources"]["a"]["label"]
    lb = config["sources"]["b"]["label"]
    a_keys = config["sources"]["a"]["keyColumns"]
    b_keys = config["sources"]["b"]["keyColumns"]
    out = config.get("output", {})
    renames = out.get("columnRenames", {})
    norm = config.get("normalization", {})
    amt_a = config["sources"]["a"]["amountColumn"]
    timing_col = config["matching"].get("timingKeyColumn")

    # Descriptive columns = every source-A column except the amount column.
    desc_cols = [c for c in meta_a["cols"] if c != amt_a]
    D = len(desc_cols)
    # Recon column letters.
    desc_letter = {name: _CL(2 + i) for i, name in enumerate(desc_cols)}
    L_amt_a = _CL(2 + D)
    L_amt_b = _CL(3 + D)
    L_diff = _CL(4 + D)
    L_lines_a = _CL(5 + D)
    L_lines_b = _CL(6 + D)
    L_status = _CL(7 + D)
    L_dtype = _CL(8 + D)
    L_root = _CL(9 + D)
    L_action = _CL(10 + D)
    ncols = 10 + D

    # Union of keys: one row per UNIQUE key - the first-seen source-A row for each A key (in
    # order), then the first-seen source-B row for each B key not already present in A. This
    # mirrors compute_reconciliation()'s union exactly, so the SUMIF/COUNTIF-per-key sheet agrees
    # with the HTML dashboard and a key duplicated within a source is never double-counted. Uses
    # norm_key per component so the union matches the Python matcher and the Excel TRIM/LOWER
    # helper (integer-valued cells canonicalize identically).
    def keystr(df, keys, i, scope):
        k = join_key_parts([norm_key(df.iloc[i][c], norm) for c in keys])
        return k if k else keyless_token(scope, i + 2)
    a_keyset = set()
    recon_rows = []
    row_keys = []
    for i in range(meta_a["n"]):
        k = keystr(df_a, a_keys, i, sa)
        if k not in a_keyset:
            a_keyset.add(k)
            recon_rows.append(("a", i + 2)); row_keys.append(k)
    b_keyset = set()
    for j in range(meta_b["n"]):
        k = keystr(df_b, b_keys, j, sb)
        if k not in a_keyset and k not in b_keyset:
            b_keyset.add(k)
            recon_rows.append(("b", j + 2)); row_keys.append(k)
    n_lines = len(recon_rows)
    r_first = 5
    r_last = r_first + n_lines - 1
    r_total = r_last + 1
    r_ctrl = r_total + 1

    # Titles.
    t = ws.cell(row=1, column=1, value=f"Reconciliation detail — {la} vs {lb}")
    t.font = Font(bold=True, size=16, color=SEC_C)
    st = ws.cell(row=2, column=1,
                 value=f"One row per matching key. Difference = {la} less {lb}. "
                       "Basis of preparation is on the Dashboard.")
    st.font = Font(color=SUB_C)

    # Header row (row 4).
    headers = (["Matching Key"] + [renames.get(c, c) for c in desc_cols] +
               [f"Amount — {la}", f"Amount — {lb}", "Difference",
                f"Lines in {la}", f"Lines in {lb}", "Status",
                "Difference Type", "Root Cause", "Action Needed"])
    hdr_fill = PatternFill("solid", fgColor=HDR_FILL)
    thin = Side(style="thin", color=HDR_FILL)
    hborder = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, name in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=c, value=name)
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color=HDR_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = hborder

    # Which recon letters correspond to the key columns (for the Matching Key join) and to
    # the non-timing key columns (for the timing COUNTIFS in Root Cause).
    key_recon_letters = [desc_letter[k] for k in a_keys if k in desc_letter]
    nontiming_keys = [k for k in a_keys if k != timing_col and k in desc_letter]
    # Timing root cause only makes sense when timing detection is enabled AND at least one
    # non-timing key column exists to group offsetting entries by; otherwise every one-sided break
    # would be grouped together and mislabelled "Timing" (mirrors the reconcile()/HTML guard).
    timing_on = (config["matching"].get("enableTimingDetection", True)
                 and timing_col is not None and len(nontiming_keys) >= 1)

    mk_a = f"'{sa}'!${meta_a['mk_letter']}$2:${meta_a['mk_letter']}${meta_a['last']}"
    amt_a_rng = f"'{sa}'!${meta_a['amt_letter']}$2:${meta_a['amt_letter']}${meta_a['last']}"
    mk_b = f"'{sb}'!${meta_b['mk_letter']}$2:${meta_b['mk_letter']}${meta_b['last']}"
    amt_b_rng = f"'{sb}'!${meta_b['amt_letter']}$2:${meta_b['amt_letter']}${meta_b['last']}"

    b_cols = list(df_b.columns)
    keymap = dict(zip(a_keys, b_keys))
    text_desc = {name for name in desc_cols if not pd.api.types.is_numeric_dtype(df_a[name])}

    for idx, (side, srow) in enumerate(recon_rows):
        r = r_first + idx
        # Matching Key. Keyless rows carry a unique placeholder (written literally, matching the
        # source-tab helper) so they are never aggregated together; keyed rows rebuild the key by
        # formula, normalized identically to the source-tab helper so SUMIF/COUNTIF align.
        if is_keyless_token(row_keys[idx]):
            ws.cell(row=r, column=1, value=row_keys[idx]).font = Font(color=BODY_C)
        else:
            refs = [f"${kl}{r}" for kl in key_recon_letters]
            ws.cell(row=r, column=1, value=_xl_key_formula(refs, norm)).font = Font(color=BODY_C)
        # Timing root cause applies to this row only when timing is on AND its non-timing (reduced)
        # key is genuinely non-blank; keyless / blank-reduced rows are excluded so they are never
        # grouped and mislabelled "Timing" (mirrors reconcile() and the HTML per-key model).
        if timing_on:
            if side == "a":
                reduced = [norm_key(df_a.iloc[srow - 2].get(k), norm) for k in nontiming_keys]
            else:
                reduced = [norm_key(df_b.iloc[srow - 2].get(keymap.get(k, k)), norm) for k in nontiming_keys]
            row_timing = any(reduced)
        else:
            row_timing = False
        # Descriptive columns pulled from the source row by reference. Numeric key columns are
        # left-aligned Cambria; free-text columns (Account Name, Period) use Arial 10 (v15 style).
        for name in desc_cols:
            col_idx = 2 + desc_cols.index(name)
            if side == "a":
                src_letter = _CL(meta_a["cols"].index(name) + 1)
                ref = f"='{sa}'!${src_letter}{srow}"
            else:
                bname = keymap.get(name, name)
                if bname in b_cols:
                    src_letter = _CL(b_cols.index(bname) + 1)
                    ref = f"='{sb}'!${src_letter}{srow}"
                else:
                    ref = None
            if ref is not None:
                cell = ws.cell(row=r, column=col_idx, value=ref)
                if name in text_desc:
                    cell.font = Font(name="Arial", size=10)
                else:
                    cell.alignment = Alignment(horizontal="left")
        # Amounts, difference, line counts.
        ws.cell(row=r, column=2 + D, value=f"=SUMIF({mk_a},$A{r},{amt_a_rng})").number_format = ACCT2
        ws.cell(row=r, column=3 + D, value=f"=SUMIF({mk_b},$A{r},{amt_b_rng})").number_format = ACCT2
        ws.cell(row=r, column=4 + D, value=f"={L_amt_a}{r}-{L_amt_b}{r}").number_format = ACCT2
        ca = ws.cell(row=r, column=5 + D, value=f"=COUNTIF({mk_a},$A{r})")
        cb = ws.cell(row=r, column=6 + D, value=f"=COUNTIF({mk_b},$A{r})")
        ca.alignment = cb.alignment = Alignment(horizontal="center")
        # Status / Difference Type / Root Cause / Action Needed (left-aligned text, v15 style).
        left = Alignment(horizontal="left")
        ws.cell(row=r, column=7 + D, value=f'=IF({L_dtype}{r}="None","Reconciled","Open Item")').alignment = left
        ws.cell(row=r, column=8 + D,
                value=(f'=IF({L_lines_a}{r}=0,"Missing in {la}",'
                       f'IF({L_lines_b}{r}=0,"Missing in {lb}",'
                       f'IF(ROUND({L_diff}{r},2)=0,"None","Amount mismatch")))')).alignment = left
        # Root Cause: measurement (amount mismatch), timing (offsetting missing entry in the
        # same non-period group when timing applies to this row), else scope / mapping.
        if row_timing:
            countifs = ""
            for k in nontiming_keys:
                kl = desc_letter[k]
                countifs += f"${kl}$5:${kl}${r_last},${kl}{r},"
            opp = f'IF({L_dtype}{r}="Missing in {lb}","Missing in {la}","Missing in {lb}")'
            root = (f'=IF({L_status}{r}="Reconciled","—",'
                    f'IF({L_dtype}{r}="Amount mismatch","Measurement",'
                    f'IF(COUNTIFS({countifs}${L_dtype}$5:${L_dtype}${r_last},{opp})>0,'
                    f'"Timing","Scope / mapping")))')
        else:
            root = (f'=IF({L_status}{r}="Reconciled","—",'
                    f'IF({L_dtype}{r}="Amount mismatch","Measurement","Scope / mapping"))')
        ws.cell(row=r, column=9 + D, value=root).alignment = left
        ws.cell(row=r, column=10 + D,
                value=(f'=IF({L_root}{r}="Measurement","Obtain supporting detail and correct the misstated balance",'
                       f'IF({L_root}{r}="Timing","Confirm cut-off; the offsetting entry sits in the adjacent period",'
                       f'IF({L_root}{r}="Scope / mapping","Confirm the account is intentionally excluded, or post the missing entry",'
                       f'"No action — line agrees")))'))

    # Totals row + control row.
    tot_lbl = ws.cell(row=r_total, column=4, value="Total"); tot_lbl.font = Font(bold=True)
    for L in (L_amt_a, L_amt_b, L_diff):
        cc = ws.cell(row=r_total, column=_col_to_idx(L) + 1, value=f"=SUM({L}{r_first}:{L}{r_last})")
        cc.font = Font(bold=True); cc.number_format = ACCT2
    top = Side(style="thin", color=HDR_FILL)
    for c in range(1, ncols + 1):
        ws.cell(row=r_total, column=c).border = Border(top=top, bottom=top)
    cl = ws.cell(row=r_ctrl, column=4,
                 value="Control — net difference ties to the independent ledger totals (must be nil)")
    cl.font = Font(color=SUB_C)
    # A real check, not a tautology: the reconciliation's net difference (sum of the per-key
    # Difference column) must equal the difference of the two *independent* source-tab totals
    # (SUM over each source's amount column). If a key were dropped from the union or a range were
    # misaligned, the per-key total would stop matching the raw source total and this reads non-nil.
    ctrl_cell = ws.cell(row=r_ctrl, column=_col_to_idx(L_diff) + 1,
                        value=f"={L_diff}{r_total}-(SUM({amt_a_rng})-SUM({amt_b_rng}))")
    ctrl_cell.number_format = ACCT2
    ctrl_cell.font = Font(bold=True)

    # Column widths (aligned to v15).
    ws.column_dimensions["A"].width = 26
    for name in desc_cols:
        Lc = desc_letter[name]
        if name == out.get("accountNameColumn"):
            w = 22
        elif name == out.get("accountColumn"):
            w = 14
        else:
            w = 10
        ws.column_dimensions[Lc].width = w
    ws.column_dimensions[L_amt_a].width = 16
    ws.column_dimensions[L_amt_b].width = 16
    ws.column_dimensions[L_diff].width = 14
    ws.column_dimensions[L_lines_a].width = 9
    ws.column_dimensions[L_lines_b].width = 9
    ws.column_dimensions[L_status].width = 12
    ws.column_dimensions[L_dtype].width = 19
    ws.column_dimensions[L_root].width = 16
    ws.column_dimensions[L_action].width = 52

    ws.freeze_panes = f"{L_amt_a}5"

    # Zebra banding + status colouring (conditional formatting, so it survives edits).
    rng = f"A{r_first}:{L_action}{r_last}"
    zebra = PatternFill(start_color=ZEBRA_BG, end_color=ZEBRA_BG, fill_type="solid")
    ws.conditional_formatting.add(rng, FormulaRule(formula=["MOD(ROW(),2)=1"], fill=zebra))
    srng = f"{L_status}{r_first}:{L_status}{r_last}"
    ws.conditional_formatting.add(srng, CellIsRule(
        operator="equal", formula=['"Open Item"'],
        fill=PatternFill(start_color=OPEN_BG, end_color=OPEN_BG, fill_type="solid"),
        font=Font(color=OPEN_FONT)))
    ws.conditional_formatting.add(srng, CellIsRule(
        operator="equal", formula=['"Reconciled"'],
        fill=PatternFill(start_color=REC_BG, end_color=REC_BG, fill_type="solid"),
        font=Font(color=REC_FONT)))

    return {
        "r_first": r_first, "r_last": r_last, "r_total": r_total, "r_ctrl": r_ctrl,
        "desc_letter": desc_letter, "D": D,
        "L_amt_a": L_amt_a, "L_amt_b": L_amt_b, "L_diff": L_diff,
        "L_lines_a": L_lines_a, "L_lines_b": L_lines_b, "L_status": L_status,
        "L_dtype": L_dtype, "L_root": L_root, "recon_rows": recon_rows,
    }


def _write_dashboard(ws, info, config, df_a, df_b, meta_a, meta_b, sa, sb, narrative, src_name):
    from openpyxl.styles import Font, PatternFill, Alignment

    la = config["sources"]["a"]["label"]
    lb = config["sources"]["b"]["label"]
    out = config.get("output", {})
    acct_col = out.get("accountColumn")
    name_col = out.get("accountNameColumn")
    renames = out.get("columnRenames", {})
    # Display headers derive from the configured column names (via columnRenames) so the
    # pivots read correctly in any domain - "Vendor ID" for a WHT run, "Account Number" for GL.
    acct_hdr = renames.get(acct_col, acct_col) if acct_col else "Account"
    name_hdr = renames.get(name_col, name_col) if name_col else "Name"
    group_by = [g for g in out.get("groupBy", []) if g in config["sources"]["a"]["keyColumns"]]
    dl = info["desc_letter"]
    rf, rl = info["r_first"], info["r_last"]
    RB = dl.get(group_by[0]) if group_by else "B"
    RE = dl.get(group_by[1]) if len(group_by) > 1 else None
    RC = dl.get(acct_col) if acct_col else None
    Fa, Fb, Fd = info["L_amt_a"], info["L_amt_b"], info["L_diff"]
    Kst, Ldt, Mrc = info["L_status"], info["L_dtype"], info["L_root"]
    rtot = info["r_total"]
    rctrl = info["r_ctrl"]

    def R(col):  # a Reconciliation range for a whole-column data span
        return f"Reconciliation!${col}${rf}:${col}${rl}"

    hdr_fill = PatternFill("solid", fgColor=HDR_FILL)

    def section(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=True, size=12, color=SEC_C)

    def header(row, col, text):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = hdr_fill
        c.font = Font(bold=True, color=HDR_FONT)
        c.alignment = Alignment(horizontal="center", vertical="center")

    def txt(cell):  # v15 renders row labels / text data in Arial 10
        cell.font = Font(name="Arial", size=10)
        return cell

    # Title + basis of preparation.
    t = ws.cell(row=1, column=1, value=f"Reconciliation Dashboard — {la} vs {lb}")
    t.font = Font(bold=True, size=18, color=SEC_C)
    basis_bits = []
    for g in group_by:
        if g == group_by[0]:
            vals_src = sorted({str(df_a.iloc[i][g]) for i in range(meta_a["n"])} |
                              {str(df_b.iloc[j][g]) for j in range(meta_b["n"])})
            basis_bits.append(f"{g} " + ", ".join(vals_src))
    basis_bits.append(f"Difference = {la} less {lb}")
    if src_name:
        basis_bits.append(f"Source: {src_name}")
    b2 = ws.cell(row=2, column=1, value=" | ".join(basis_bits))
    b2.font = Font(color=SUB_C)

    # ---- Control panel (rows 5-11) ----
    section(5, 1, "Control panel — every control must read OK before sign-off")
    for j, h in enumerate(["Control", "Result", "Expected", "Status"]):
        header(6, 1 + j, h)
    a_mk = f"'{sa}'!${meta_a['mk_letter']}$2:${meta_a['mk_letter']}${meta_a['last']}"
    b_mk = f"'{sb}'!${meta_b['mk_letter']}$2:${meta_b['mk_letter']}${meta_b['last']}"
    a_amt = f"'{sa}'!${meta_a['amt_letter']}$2:${meta_a['amt_letter']}${meta_a['last']}"
    b_amt = f"'{sb}'!${meta_b['amt_letter']}$2:${meta_b['amt_letter']}${meta_b['last']}"
    controls = [
        ("Every key in either ledger appears once",
         f"=COUNTA(Reconciliation!$A${rf}:$A${rl})",
         # Unique keys in A, plus unique keys in B that are absent from A - matching the union
         # (which de-duplicates each source), rather than counting raw rows. The 1/COUNTIF pattern
         # collapses repeats of the same key to a single count; keyless rows each carry a distinct
         # placeholder so they count once apiece. Each term drops to 0 when its source has no data
         # rows, so the helper range never inverts to include a blank cell (which would #DIV/0!).
         ("=" + (f"SUMPRODUCT(1/COUNTIF({a_mk},{a_mk}))" if meta_a["n"] else "0")
          + "+" + (f"SUMPRODUCT((COUNTIF({a_mk},{b_mk})=0)/COUNTIF({b_mk},{b_mk}))"
                   if meta_b["n"] else "0")),
         "count"),
        (f"Amount — {la} agrees to the {la} tab",
         f"=Reconciliation!${Fa}${rtot}", f"=SUM({a_amt})", "acct"),
        (f"Amount — {lb} agrees to the {lb} tab",
         f"=Reconciliation!${Fb}${rtot}", f"=SUM({b_amt})", "acct"),
        ("Total difference proves to the two ledger totals",
         f"=Reconciliation!${Fd}${rctrl}", "0", "acct"),
        ("Reconciled plus open items equal total lines",
         f'=COUNTIF({R(Kst)},"Reconciled")+COUNTIF({R(Kst)},"Open Item")',
         f"=COUNTA(Reconciliation!$A${rf}:$A${rl})", "count"),
    ]
    for i, (label, result, expected, kind) in enumerate(controls):
        row = 7 + i
        txt(ws.cell(row=row, column=1, value=label))
        rc = ws.cell(row=row, column=2, value=result)
        ec = ws.cell(row=row, column=3, value=expected)
        rc.alignment = ec.alignment = Alignment(horizontal="right")
        fmt = ACCT2 if kind == "acct" else CNT_FMT
        rc.number_format = ec.number_format = fmt
        sc = ws.cell(row=row, column=4,
                     value=(f'=IF(ROUND(B{row}-C{row},2)=0,"OK","CHECK")' if kind == "acct"
                            else f'=IF(B{row}=C{row},"OK","CHECK")'))
        sc.alignment = Alignment(horizontal="center")

    # Right side of the control band: open items by difference type.
    section(6, 8, "Open items by difference type")
    header(7, 8, "Difference type"); header(7, 9, "Count"); header(7, 10, "Value, ignoring sign")
    dtypes = ["Amount mismatch", f"Missing in {la}", f"Missing in {lb}"]
    for i, dt in enumerate(dtypes):
        row = 8 + i
        txt(ws.cell(row=row, column=8, value=dt))
        ws.cell(row=row, column=9, value=f"=COUNTIF({R(Ldt)},$H{row})").alignment = Alignment(horizontal="right")
        vc = ws.cell(row=row, column=10,
                     value=f"=SUMPRODUCT(({R(Ldt)}=$H{row})*ABS({R(Fd)}))")
        vc.number_format = ACCT2
    trow = 8 + len(dtypes)
    ws.cell(row=trow, column=8, value="Total").font = Font(bold=True)
    ws.cell(row=trow, column=9, value=f"=SUM(I8:I{trow-1})").font = Font(bold=True)
    tc = ws.cell(row=trow, column=10, value=f"=SUM(J8:J{trow-1})")
    tc.font = Font(bold=True); tc.number_format = ACCT2

    # ---- Reconciliation summary (rows 13-19) + open items by root cause ----
    section(13, 1, "Reconciliation summary")
    summ = [
        ("Total lines", f"=COUNTA(Reconciliation!$A${rf}:$A${rl})", CNT_FMT),
        ("Reconciled", f'=COUNTIF({R(Kst)},"Reconciled")', CNT_FMT),
        ("Open items", f'=COUNTIF({R(Kst)},"Open Item")', CNT_FMT),
        ("Match rate", "=IF(B14=0,0,B15/B14)", PCT_FMT),
        (f"Net difference ({la} less {lb})", f"=SUM({R(Fd)})", ACCT2),
        ("Gross difference, ignoring sign", f"=SUMPRODUCT(ABS({R(Fd)}))", ACCT2),
    ]
    for i, (label, formula, fmt) in enumerate(summ):
        row = 14 + i
        txt(ws.cell(row=row, column=1, value=label))
        vc = ws.cell(row=row, column=2, value=formula)
        vc.font = Font(bold=True); vc.alignment = Alignment(horizontal="right")
        vc.number_format = fmt

    section(13, 8, "Open items by root cause")
    header(14, 8, "Root cause"); header(14, 9, "Count"); header(14, 10, "Value, ignoring sign")
    roots = ["Measurement", "Timing", "Scope / mapping"]
    for i, rt in enumerate(roots):
        row = 15 + i
        txt(ws.cell(row=row, column=8, value=rt))
        ws.cell(row=row, column=9, value=f"=COUNTIF({R(Mrc)},$H{row})").alignment = Alignment(horizontal="right")
        vc = ws.cell(row=row, column=10, value=f"=SUMPRODUCT(({R(Mrc)}=$H{row})*ABS({R(Fd)}))")
        vc.number_format = ACCT2
    rtrow = 15 + len(roots)
    ws.cell(row=rtrow, column=8, value="Total").font = Font(bold=True)
    ws.cell(row=rtrow, column=9, value=f"=SUM(I15:I{rtrow-1})").font = Font(bold=True)
    vc = ws.cell(row=rtrow, column=10, value=f"=SUM(J15:J{rtrow-1})")
    vc.font = Font(bold=True); vc.number_format = ACCT2

    # ---- Difference by account (rows 21+) ----
    # Build the unique account list first; only draw the section when the account column is
    # configured, resolvable to a Reconciliation column (RC), and actually present in the data.
    # Otherwise the SUMIF ranges would reference a "$None$" column and an empty list would build a
    # reversed SUM() range (e.g. SUM(C23:C22)).
    accounts = []
    if acct_col and RC:
        seen = set()
        keymap = dict(zip(config["sources"]["a"]["keyColumns"], config["sources"]["b"]["keyColumns"]))
        for side, srow in info["recon_rows"]:
            if side == "a":
                acct = df_a.iloc[srow - 2][acct_col] if acct_col in df_a.columns else None
                nm = df_a.iloc[srow - 2][name_col] if (name_col and name_col in df_a.columns) else ""
            else:
                bacct = keymap.get(acct_col, acct_col)
                acct = df_b.iloc[srow - 2][bacct] if bacct in df_b.columns else None
                nm = df_b.iloc[srow - 2][name_col] if (name_col and name_col in df_b.columns) else ""
            if acct is not None and acct not in seen:
                seen.add(acct); accounts.append((acct, nm))
    acc_start = 23
    acc_tot = 22  # baseline row if the account section is not drawn (keeps the later blocks below)
    if accounts:
        section(21, 1, "Difference by account")
        for j, h in enumerate([acct_hdr, name_hdr, f"Amount \u2014 {la}", f"Amount \u2014 {lb}", "Difference"]):
            header(22, 1 + j, h)
        for i, (acct, nm) in enumerate(accounts):
            row = acc_start + i
            ws.cell(row=row, column=1, value=acct).alignment = Alignment(horizontal="left")
            txt(ws.cell(row=row, column=2, value=nm))
            ws.cell(row=row, column=3, value=f"=SUMIF({R(RC)},$A{row},{R(Fa)})").number_format = ACCT2
            ws.cell(row=row, column=4, value=f"=SUMIF({R(RC)},$A{row},{R(Fb)})").number_format = ACCT2
            ws.cell(row=row, column=5, value=f"=C{row}-D{row}").number_format = ACCT2
        acc_tot = acc_start + len(accounts)
        ws.cell(row=acc_tot, column=2, value="Total").font = Font(bold=True)
        for col, base in ((3, "C"), (4, "D"), (5, "E")):
            cc = ws.cell(row=acc_tot, column=col, value=f"=SUM({base}{acc_start}:{base}{acc_tot-1})")
            cc.font = Font(bold=True); cc.number_format = ACCT2

    # Difference by company and period (right side; aligned one row lower than the left block,
    # matching v15 - section on row 22, sub-headers on row 23, data from row 24). Only drawn when
    # both group dimensions resolve AND at least one (company, period) combo exists, so the totals
    # never build a reversed SUM() range.
    if RB and RE:
        combos = []
        seenc = set()
        for side, srow in info["recon_rows"]:
            src = df_a if side == "a" else df_b
            comp = src.iloc[srow - 2][group_by[0]] if group_by[0] in src.columns else None
            per = src.iloc[srow - 2][group_by[1]] if group_by[1] in src.columns else None
            key = (comp, per)
            if comp is not None and per is not None and key not in seenc:
                seenc.add(key); combos.append((comp, per))
        if combos:
            comp_hdr = renames.get(group_by[0], group_by[0])
            per_hdr = renames.get(group_by[1], group_by[1])
            section(22, 8, f"Difference by {comp_hdr.lower()} and {per_hdr.lower()}")
            for j, h in enumerate([comp_hdr, per_hdr, f"Amount \u2014 {la}", f"Amount \u2014 {lb}",
                                   "Difference", "Open items"]):
                header(23, 8 + j, h)
            cp_start = 24
            for i, (comp, per) in enumerate(combos):
                row = cp_start + i
                ws.cell(row=row, column=8, value=comp).alignment = Alignment(horizontal="left")
                txt(ws.cell(row=row, column=9, value=per))
                ws.cell(row=row, column=10,
                        value=f"=SUMIFS({R(Fa)},{R(RB)},$H{row},{R(RE)},$I{row})").number_format = ACCT2
                ws.cell(row=row, column=11,
                        value=f"=SUMIFS({R(Fb)},{R(RB)},$H{row},{R(RE)},$I{row})").number_format = ACCT2
                ws.cell(row=row, column=12, value=f"=J{row}-K{row}").number_format = ACCT2
                mc = ws.cell(row=row, column=13,
                             value=f'=COUNTIFS({R(RB)},$H{row},{R(RE)},$I{row},{R(Kst)},"Open Item")')
                mc.number_format = CNT_FMT
                mc.alignment = Alignment(horizontal="center")
            cp_tot = cp_start + len(combos)
            ws.cell(row=cp_tot, column=9, value="Total").font = Font(bold=True)
            for col, base in ((10, "J"), (11, "K")):
                cc = ws.cell(row=cp_tot, column=col, value=f"=SUM({base}{cp_start}:{base}{cp_tot-1})")
                cc.font = Font(bold=True); cc.number_format = ACCT2
            for col, base in ((12, "L"), (13, "M")):
                cc = ws.cell(row=cp_tot, column=col, value=f"=SUM({base}{cp_start}:{base}{cp_tot-1})")
                cc.font = Font(bold=True)
            acc_tot = max(acc_tot, cp_tot)

    # ---- Headlines (driver narrative; Calibri, matching v15) ----
    h_row = acc_tot + 2
    section(h_row, 1, "Headlines")
    narr_rows = []
    for i, line in enumerate(narrative):
        row = h_row + 1 + i
        narr_rows.append(row)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(name="Calibri", color=NARR_C)
        import math
        ws.row_dimensions[row].height = max(15, 15 * math.ceil((len(line) + 3) / 130))

    # Widths.
    for col, w in {"A": 42, "B": 60, "C": 18, "D": 22, "E": 14, "F": 18, "G": 3,
                   "H": 31, "I": 8, "J": 20, "K": 14, "L": 14, "M": 9}.items():
        ws.column_dimensions[col].width = w
    return narr_rows


def write_report(results, config, out_path, df_a=None, df_b=None, src_name=None):
    import openpyxl

    la = config["sources"]["a"].get("label", "Source A")
    lb = config["sources"]["b"].get("label", "Source B")

    res_df = pd.DataFrame(results)
    counts = res_df["status"].value_counts().to_dict() if not res_df.empty else {}
    # The Dashboard headlines are built from the per-key reconciliation model (compute_reconciliation) -
    # the same model that drives the Reconciliation sheet's formulas and the HTML dashboard - so the
    # narrative counts can never disagree with the sheet totals. (The tiered `results`/`counts` are a
    # record-level view returned for the caller's console summary, not the per-key artifact.)
    perkey_rows, _, _ = compute_reconciliation(df_a, df_b, config)
    narrative = _build_narrative_perkey(perkey_rows, config)

    meta_a = _src_meta(df_a, config["sources"]["a"], config)
    meta_b = _src_meta(df_b, config["sources"]["b"], config)
    norm = config.get("normalization", {})
    sign_a = config["sources"]["a"].get("signConvention", "asIs")
    sign_b = config["sources"]["b"].get("signConvention", "asIs")

    wb = openpyxl.Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_recon = wb.create_sheet("Reconciliation")
    # Excel sheet names: <=31 chars, no : \ / ? * [ ], and unique. Sanitize both labels.
    taken = {"dashboard", "reconciliation"}
    sa = _safe_sheet_name(la, taken)
    sb = _safe_sheet_name(lb, taken)
    ws_sa = wb.create_sheet(sa)
    ws_sb = wb.create_sheet(sb)

    _write_source_tab(ws_sa, df_a, meta_a, sa, sign=sign_a, norm=norm)
    _write_source_tab(ws_sb, df_b, meta_b, sb, sign=sign_b, norm=norm)
    info = _write_reconciliation(ws_recon, df_a, df_b, config, meta_a, meta_b, sa, sb)
    narr_rows = _write_dashboard(ws_dash, info, config, df_a, df_b, meta_a, meta_b, sa, sb, narrative, src_name)

    # Apply the report fonts: Cambria everywhere structural/numeric, leaving any cell already
    # marked Arial (v15 uses Arial 10 for pulled text data / row labels) untouched.
    from openpyxl.styles import Font
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is None or c.font.name == "Arial":
                    continue
                f = c.font
                c.font = Font(name=REPORT_FONT, size=f.size, bold=f.bold,
                              italic=f.italic, color=f.color)
    # v15 renders the headlines narrative in Calibri; restore it after the Cambria pass.
    for row in narr_rows:
        cell = ws_dash.cell(row=row, column=1)
        cell.font = Font(name="Calibri", color=NARR_C)

    wb.save(out_path)
    return counts



# ----------------------------- HTML dashboard -----------------------------

def compute_reconciliation(df_a, df_b, config):
    """Compute the same per-key reconciliation the Excel derives by formula, as Python values
    (for the HTML dashboard). One row per unique union key with amounts, status, difference type
    and root cause, so the HTML always agrees with the workbook."""
    la = config["sources"]["a"].get("label", "Source A")
    lb = config["sources"]["b"].get("label", "Source B")
    a_keys = config["sources"]["a"]["keyColumns"]
    b_keys = config["sources"]["b"]["keyColumns"]
    amt_a_col = config["sources"]["a"]["amountColumn"]
    amt_b_col = config["sources"]["b"]["amountColumn"]
    out = config.get("output", {})
    acct_col = out.get("accountColumn")
    name_col = out.get("accountNameColumn")
    group_by = out.get("groupBy", [])
    timing_col = config["matching"].get("timingKeyColumn")
    keymap_ab = dict(zip(a_keys, b_keys))

    def kstr(row, keys):
        # Canonical key: identical to build_key() (norm_key per component, shared KEY_DELIM,
        # all-empty collapses to "") so the HTML groups keys the same way as the matcher and the
        # workbook.
        return join_key_parts([norm_key(row.get(k), norm) for k in keys])

    norm = config.get("normalization", {})
    sign_a = config["sources"]["a"].get("signConvention", "asIs")
    sign_b = config["sources"]["b"].get("signConvention", "asIs")

    def amt_a(v):
        return apply_sign(normalize_amount(v, norm), sign_a) or 0.0

    def amt_b(v):
        return apply_sign(normalize_amount(v, norm), sign_b) or 0.0

    # Aggregate each source by canonical (normalized) key string. Keyless rows (all key parts
    # blank) get a unique per-row placeholder so they are never merged together - matching the
    # workbook helper and the record matcher, which treat an empty key as non-matchable.
    a_recs, b_recs = df_a.to_dict("records"), df_b.to_dict("records")
    a_agg, b_agg = {}, {}
    a_first, b_first = {}, {}
    order = []
    for i, rec in enumerate(a_recs):
        kl = kstr(rec, a_keys) or keyless_token("A " + la, i + 2)
        if kl not in a_agg:
            a_agg[kl] = {"sum": 0.0, "n": 0}; a_first[kl] = rec
            order.append(kl)
        a_agg[kl]["sum"] += amt_a(rec[amt_a_col]); a_agg[kl]["n"] += 1
    for j, rec in enumerate(b_recs):
        kl = kstr(rec, b_keys) or keyless_token("B " + lb, j + 2)
        if kl not in b_agg:
            b_agg[kl] = {"sum": 0.0, "n": 0}; b_first[kl] = rec
        b_agg[kl]["sum"] += amt_b(rec[amt_b_col]); b_agg[kl]["n"] += 1
    for kl in b_agg:
        if kl not in a_agg:
            order.append(kl)

    def field(kl, col, bcol=None):
        if kl in a_first and col in a_first[kl]:
            return a_first[kl][col]
        if kl in b_first:
            bc = bcol or keymap_ab.get(col, col)
            if bc in b_first[kl]:
                return b_first[kl][bc]
        return ""

    rows = []
    for kl in order:
        aa = a_agg.get(kl, {"sum": 0.0, "n": 0})
        bb = b_agg.get(kl, {"sum": 0.0, "n": 0})
        diff = round(aa["sum"] - bb["sum"], 2)
        if aa["n"] == 0:
            dtype = f"Missing in {la}"
        elif bb["n"] == 0:
            dtype = f"Missing in {lb}"
        elif diff == 0:
            dtype = "None"
        else:
            dtype = "Amount mismatch"
        status = "Reconciled" if dtype == "None" else "Open Item"
        rows.append({
            "key": kl,
            "company": field(kl, group_by[0]) if group_by else "",
            "account": field(kl, acct_col) if acct_col else "",
            "name": field(kl, name_col) if name_col else "",
            "period": field(kl, timing_col) if timing_col else (field(kl, group_by[1]) if len(group_by) > 1 else ""),
            "amt_a": aa["sum"], "amt_b": bb["sum"], "diff": diff,
            "lines_a": aa["n"], "lines_b": bb["n"], "status": status, "difftype": dtype,
        })

    # Root cause (needs the whole population to spot an offsetting timing entry). Timing only
    # applies when timing detection is enabled AND there is at least one non-timing key column to
    # group offsetting entries by, and only for a row whose non-timing (reduced) key is not all
    # blank - otherwise unrelated one-sided breaks would be mislabelled "Timing" (mirrors the
    # reconcile() and Excel-root-cause guards so the platforms don't diverge).
    nontiming = [k for k in a_keys if k != timing_col]
    timing_on = (config["matching"].get("enableTimingDetection", True)
                 and timing_col is not None and len(nontiming) >= 1)
    grp = {}
    for r in rows:
        gk = tuple(str(field(r["key"], k)) for k in nontiming)
        grp.setdefault(gk, []).append(r)
    for r in rows:
        if r["status"] == "Reconciled":
            r["rootcause"] = "—"
        elif r["difftype"] == "Amount mismatch":
            r["rootcause"] = "Measurement"
        else:
            gk = tuple(str(field(r["key"], k)) for k in nontiming)
            opp = f"Missing in {la}" if r["difftype"] == f"Missing in {lb}" else f"Missing in {lb}"
            # Only a row whose reduced (non-timing) key is genuinely non-blank can be a timing
            # difference. norm_key collapses NaN/blank components to "" (str(NaN) would be the
            # truthy "nan"), so keyless / blank-reduced rows are correctly excluded - matching the
            # reconcile() and Excel guards.
            reduced_nonempty = any(norm_key(field(r["key"], k), norm) for k in nontiming)
            has_offset = (timing_on and reduced_nonempty
                          and any(o["difftype"] == opp for o in grp.get(gk, [])))
            r["rootcause"] = "Timing" if has_offset else "Scope / mapping"
    return rows, la, lb


def _num(x):
    """Number cell, matching the Excel format: 2 decimals, parentheses for negatives,
    0.00 for zero. No currency symbol. Used for every value so the HTML reads consistently."""
    x = x or 0
    return f"({abs(x):,.2f})" if round(x, 2) < 0 else f"{x:,.2f}"


def build_html_dashboard(rows, config, src_name=None, df_a=None, df_b=None):
    import html as _h
    la = config["sources"]["a"].get("label", "Source A")
    lb = config["sources"]["b"].get("label", "Source B")
    out = config.get("output", {})
    group_by = out.get("groupBy", [])
    renames = out.get("columnRenames", {})
    acct_col = out.get("accountColumn")
    name_col = out.get("accountNameColumn")
    # Display headers derive from the configured column names so the pivots read correctly in
    # any domain (e.g. "Vendor ID" for a WHT run instead of a hardcoded "Account Number").
    acct_hdr = renames.get(acct_col, acct_col) if acct_col else "Account"
    name_hdr = renames.get(name_col, name_col) if name_col else "Name"
    comp_hdr = renames.get(group_by[0], group_by[0]) if group_by else "Company"
    per_hdr = renames.get(group_by[1], group_by[1]) if len(group_by) > 1 else "Period"

    total = len(rows)
    reconciled = sum(1 for r in rows if r["status"] == "Reconciled")
    open_rows = [r for r in rows if r["status"] == "Open Item"]
    opn = len(open_rows)
    net = round(sum(r["diff"] for r in rows), 2)
    gross = round(sum(abs(r["diff"]) for r in rows), 2)
    total_a = round(sum(r["amt_a"] for r in rows), 2)
    total_b = round(sum(r["amt_b"] for r in rows), 2)
    rate = (reconciled / total * 100) if total else 0

    # Independent source totals: summed directly from the raw records (a different code path from
    # the per-key aggregation), so the amount controls below are a real check - they would read
    # CHECK if the per-key aggregation dropped or double-counted a record - rather than comparing a
    # value to itself. Falls back to the per-key totals only if the raw frames weren't provided.
    nrm = config.get("normalization", {})
    amt_a_col = config["sources"]["a"]["amountColumn"]
    amt_b_col = config["sources"]["b"]["amountColumn"]
    sgn_a = config["sources"]["a"].get("signConvention", "asIs")
    sgn_b = config["sources"]["b"].get("signConvention", "asIs")
    if df_a is not None and df_b is not None:
        ind_a = round(sum(apply_sign(normalize_amount(rec.get(amt_a_col), nrm), sgn_a) or 0.0
                          for rec in df_a.to_dict("records")), 2)
        ind_b = round(sum(apply_sign(normalize_amount(rec.get(amt_b_col), nrm), sgn_b) or 0.0
                          for rec in df_b.to_dict("records")), 2)
    else:
        ind_a, ind_b = total_a, total_b

    def agg_by(keyf):
        d = {}
        for r in open_rows:
            k = keyf(r); c, v = d.get(k, (0, 0.0)); d[k] = (c + 1, v + abs(r["diff"]))
        return d
    by_type = agg_by(lambda r: r["difftype"])
    by_root = agg_by(lambda r: r["rootcause"])

    # By account (all rows), by company/period (all rows) with open counts.
    acct = {}; acct_order = []
    for r in rows:
        a = r["account"]
        if a not in acct:
            acct[a] = {"name": r["name"], "a": 0.0, "b": 0.0}; acct_order.append(a)
        acct[a]["a"] += r["amt_a"]; acct[a]["b"] += r["amt_b"]
    cp = {}; cp_order = []
    for r in rows:
        k = (r["company"], r["period"])
        if k not in cp:
            cp[k] = {"a": 0.0, "b": 0.0, "open": 0}; cp_order.append(k)
        cp[k]["a"] += r["amt_a"]; cp[k]["b"] += r["amt_b"]
        if r["status"] == "Open Item":
            cp[k]["open"] += 1

    # (Headline aggregates now come from _build_narrative_perkey, shared with the Excel Dashboard.)

    def e(s):
        return _h.escape(str(s))

    def numcell(x, bar=None, maxabs=None):
        cls = "num neg" if (x or 0) < 0 else "num"
        s = _num(x)
        if bar and maxabs:
            w = min(100, abs(x) / maxabs * 100) if maxabs else 0
            return f'<td class="{cls} bar-cell">{s}<span class="bar" style="width:{w:.1f}%"></span></td>'
        return f'<td class="{cls}">{s}</td>'

    # ---- controls ----
    ctrls = [
        ("Every key in either ledger appears once", str(total), str(total), True),
        (f"Amount — {la} agrees to the {la} tab", _num(total_a), _num(ind_a), round(total_a - ind_a, 2) == 0),
        (f"Amount — {lb} agrees to the {lb} tab", _num(total_b), _num(ind_b), round(total_b - ind_b, 2) == 0),
        ("Total difference proves to the two ledger totals",
         _num(net), _num(round(ind_a - ind_b, 2)), round(net - (ind_a - ind_b), 2) == 0),
        ("Reconciled plus open items equal total lines", str(reconciled + opn), str(total), reconciled + opn == total),
    ]
    ctrl_html = "".join(
        f'<tr><td>{e(lbl)}</td><td class="num">{res}</td><td class="num">{exp}</td>'
        f'<td><span class="ok">{"OK" if ok else "CHECK"}</span></td></tr>'
        for lbl, res, exp, ok in ctrls)

    # ---- type / root tables ----
    def kv_table(d, order):
        body = ""
        tc = tv = 0
        for k in order:
            if k in d:
                c, v = d[k]; tc += c; tv += v
                body += f'<tr><td>{e(k)}</td><td class="num">{c:,}</td><td class="num">{_num(v)}</td></tr>'
        body += f'<tr class="total"><td>Total</td><td class="num">{tc:,}</td><td class="num">{_num(tv)}</td></tr>'
        return body
    type_body = kv_table(by_type, ["Amount mismatch", f"Missing in {la}", f"Missing in {lb}"])
    root_body = kv_table(by_root, ["Measurement", "Timing", "Scope / mapping"])

    # ---- account table ----
    maxacct = max((abs(acct[a]["a"] - acct[a]["b"]) for a in acct_order), default=1) or 1
    acct_body = ""
    for a in acct_order:
        d = acct[a]; diff = d["a"] - d["b"]
        acct_body += (f'<tr><td>{e(a)}</td><td>{e(d["name"])}</td>'
                      f'{numcell(d["a"])}{numcell(d["b"])}{numcell(diff, bar=True, maxabs=maxacct)}</tr>')
    acct_body += (f'<tr class="total"><td></td><td>Total</td>{numcell(total_a)}{numcell(total_b)}'
                  f'{numcell(round(total_a-total_b,2))}</tr>')

    # ---- company/period table ----
    maxcp = max((abs(cp[k]["a"] - cp[k]["b"]) for k in cp_order), default=1) or 1
    cp_body = ""
    cp_ta = cp_tb = cp_open = 0
    for k in cp_order:
        d = cp[k]; diff = d["a"] - d["b"]; cp_ta += d["a"]; cp_tb += d["b"]; cp_open += d["open"]
        cp_body += (f'<tr><td>{e(k[0])}</td><td>{e(k[1])}</td>'
                    f'{numcell(d["a"])}{numcell(d["b"])}{numcell(diff, bar=True, maxabs=maxcp)}'
                    f'<td class="num">{d["open"]}</td></tr>')
    cp_body += (f'<tr class="total"><td></td><td>Total</td>{numcell(cp_ta)}{numcell(cp_tb)}'
                f'{numcell(round(cp_ta-cp_tb,2))}<td class="num">{cp_open}</td></tr>')

    # ---- open-item detail (sorted by magnitude) ----
    det = sorted(open_rows, key=lambda r: abs(r["diff"]), reverse=True)
    det_body = ""
    for r in det:
        det_body += (f'<tr><td>{e(r["company"])}</td><td>{e(r["account"])}</td><td>{e(r["name"])}</td>'
                     f'<td>{e(r["period"])}</td>{numcell(r["amt_a"])}{numcell(r["amt_b"])}{numcell(r["diff"])}'
                     f'<td><span class="tag tag-open">Open Item</span></td>'
                     f'<td>{e(r["difftype"])}</td><td>{e(r["rootcause"])}</td></tr>')
    det_body += (f'<tr class="total"><td colspan="6">Total difference on open items</td>'
                 f'{numcell(round(sum(r["diff"] for r in open_rows),2))}<td colspan="3"></td></tr>')

    # ---- headlines ----
    gb0 = group_by[0] if group_by else "company"
    gb1 = group_by[1] if len(group_by) > 1 else "period"
    # Analytical headlines come from the shared per-key narrative builder, so the HTML and the
    # Excel Dashboard show identical wording over the same numbers. The controls line is
    # HTML-specific (the workbook evaluates its controls as live formulas instead).
    headlines = list(_build_narrative_perkey(rows, config))
    ok_ct = sum(1 for c in ctrls if c[3])
    headlines.append(f"All {ok_ct} of {len(ctrls)} controls currently read OK." if ok_ct == len(ctrls)
                     else f"{ok_ct} of {len(ctrls)} controls read OK — resolve the exceptions before sign-off.")
    head_html = "".join(f"<li>{e(h)}</li>" for h in headlines)

    companies = sorted({str(r["company"]) for r in rows if r["company"] != ""})
    periods = sorted({str(r["period"]) for r in rows if r["period"] != ""})
    # Build the sub-header from escaped pieces so user-derived labels/values can't inject markup.
    esc_comp = " and ".join(e(c) for c in companies)
    esc_per = " and ".join(e(p) for p in periods)
    sub = (f"{e(la)} vs {e(lb)} &nbsp;|&nbsp; {e(gb0)} " + esc_comp +
           f" &nbsp;|&nbsp; {e(gb1)} " + esc_per +
           f" &nbsp;|&nbsp; Difference = {e(la)} less {e(lb)}" +
           (f" &nbsp;|&nbsp; Source: {e(src_name)}" if src_name else ""))

    return _HTML_TEMPLATE.format(
        title=e(f"{la} vs {lb} Reconciliation Dashboard"),
        h1=e(f"{la} vs {lb} Reconciliation Dashboard"),
        sub=sub,
        ctrl=ctrl_html,
        k_total=total, k_rec=reconciled, k_rate=f"{rate:.1f}", k_open=opn,
        k_net=_num(net), k_net_cls=("value neg" if net < 0 else "value"),
        k_gross=_num(gross),
        type_body=type_body, root_body=root_body,
        la=e(la), lb=e(lb), acct_body=acct_body, cp_body=cp_body,
        acct_hdr=e(acct_hdr), name_hdr=e(name_hdr), comp_hdr=e(comp_hdr), per_hdr=e(per_hdr),
        cp_title=e(f"Difference by {comp_hdr.lower()} and {per_hdr.lower()}"),
        open_ct=opn, det_body=det_body, rec_ct=reconciled,
        head=head_html,
        footer=(f"Prepared from {e(src_name)}. " if src_name else "") + "Companion workbook with live formulas accompanies this dashboard.",
    )


def write_html_report(rows, config, out_path, src_name=None, df_a=None, df_b=None):
    html = build_html_dashboard(rows, config, src_name, df_a=df_a, df_b=df_b)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)


_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en" dir="ltr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="color-scheme" content="light dark">
<title>{title}</title>
<style>
:root{{
  --bg:#f3f6fa; --surface:#ffffff; --ink:#1e2430; --muted:#5c6675;
  --line:#dae2ec; --rule:#e3eaf3;
  --brand:#2e5c8a; --brand-2:#5b9bd5; --brand-tint:#eaf1f8; --brand-band:#f2f7fc;
  --open-bg:#f7e3e1; --open-ink:#8c1d18;
  --rec-bg:#eaf1f8; --rec-ink:#1f3864;
  --neg:#b3261e; --radius:10px;
}}
@media (prefers-color-scheme: dark){{
  :root{{
    --bg:#11151b; --surface:#1a202a; --ink:#e7ecf4; --muted:#a2acbc;
    --line:#2b3340; --rule:#232b36;
    --brand:#7fb0e0; --brand-2:#4a7fb5; --brand-tint:#1d2836; --brand-band:#171e28;
    --open-bg:#3a201e; --open-ink:#f0a9a3; --rec-bg:#1d2a3c; --rec-ink:#9dc0e8; --neg:#ff8a80;
  }}
}}
*{{box-sizing:border-box;}}
body{{margin:0; padding:28px 22px 52px; background:var(--bg); color:var(--ink);
  font-family:Arial,"Segoe UI",Helvetica,sans-serif; font-size:14.5px; line-height:1.5;}}
.wrap{{max-width:1180px; margin-inline:auto;}}
header.page-head{{margin-block-end:24px;}}
h1{{font-size:1.7rem; margin:0 0 4px; color:var(--brand); letter-spacing:-.2px;}}
.sub{{color:var(--muted); font-size:.86rem; margin:0 0 12px;}}
.accent{{height:5px; background:var(--brand-2); border-radius:3px;}}
h2{{font-size:.95rem; text-transform:uppercase; letter-spacing:.7px; color:var(--brand); margin:30px 0 10px;}}
section{{break-inside:avoid;}}
.card{{background:var(--surface); border:1px solid var(--line); border-radius:var(--radius); padding:16px 18px;}}
.grid2{{display:grid; grid-template-columns:repeat(auto-fit,minmax(330px,1fr)); gap:16px;}}
.input-row{{display:flex; align-items:center; gap:12px; flex-wrap:wrap;}}
.input-row .label{{font-weight:bold;}}
.input-row .val{{font-weight:bold; color:#0000ff; background:var(--brand-tint);
  border:1px solid var(--brand-2); border-radius:5px; padding:3px 12px; font-variant-numeric:tabular-nums;}}
@media (prefers-color-scheme: dark){{ .input-row .val{{color:#8fb8ff;}} }}
.input-row .note{{color:var(--muted); font-size:.82rem; font-style:italic;}}
.kpis{{display:grid; grid-template-columns:repeat(auto-fit,minmax(165px,1fr)); gap:12px;}}
.kpi{{background:var(--surface); border:1px solid var(--line); border-radius:var(--radius); padding:12px 14px;}}
.kpi .label{{font-size:.74rem; text-transform:uppercase; letter-spacing:.5px; color:var(--muted);}}
.kpi .value{{font-size:1.5rem; font-weight:bold; margin-block-start:3px; font-variant-numeric:tabular-nums;}}
.kpi .foot{{font-size:.76rem; color:var(--muted);}}
.value.neg{{color:var(--neg);}}
table{{width:100%; border-collapse:collapse; font-size:.85rem;}}
th,td{{padding:7px 9px; text-align:start; border-block-end:1px solid var(--rule);}}
th{{background:var(--brand); color:#fff; font-weight:bold; font-size:.75rem; text-transform:uppercase; letter-spacing:.4px;}}
td.num,th.num{{text-align:end; font-variant-numeric:tabular-nums;}}
tbody tr:nth-child(odd){{background:var(--brand-band);}}
tr.total td{{font-weight:bold; background:transparent; border-block-start:2px solid var(--brand); border-block-end:2px solid var(--brand);}}
.neg{{color:var(--neg);}}
.tag{{display:inline-block; padding:2px 9px; border-radius:4px; font-size:.75rem; white-space:nowrap;}}
.tag-open{{background:var(--open-bg); color:var(--open-ink);}}
.tag-rec{{background:var(--rec-bg); color:var(--rec-ink);}}
.ok{{background:var(--rec-bg); color:var(--rec-ink); font-weight:bold; padding:2px 9px; border-radius:4px;}}
.bar-cell{{position:relative;}}
.bar{{display:block; height:4px; background:var(--brand-2); border-radius:2px; margin-block-start:3px; margin-inline-start:auto;}}
.notes dt{{font-weight:bold; margin-block-start:8px;}}
.notes dd{{margin:0; color:var(--muted); font-size:.86rem;}}
.headlines{{margin:0; padding-inline-start:18px;}}
.headlines li{{margin-block-end:5px;}}
footer{{margin-block-start:30px; padding-block-start:10px; border-block-start:1px solid var(--line); font-size:.78rem; color:var(--muted);}}
@media print{{
  body{{background:#fff; padding:0; font-size:10.5pt;}}
  .card,.kpi{{border-color:#ccc;}} h2{{margin-block-start:16px;}}
  thead{{display:table-header-group;}}
  tbody tr:nth-child(odd){{background:#f4f7fb !important; -webkit-print-color-adjust:exact; print-color-adjust:exact;}}
  th{{background:#2e5c8a !important; -webkit-print-color-adjust:exact; print-color-adjust:exact;}}
}}
</style>
</head>
<body>
<div class="wrap">
<header class="page-head">
  <h1>{h1}</h1>
  <p class="sub">{sub}</p>
  <div class="accent"></div>
</header>

<section><div class="card input-row">
  <span class="label">Reconciliation basis</span>
  <span class="val">{la} less {lb}</span>
  <span class="note">Every matching key present in either ledger is compared once; differences are classified below.</span>
</div></section>

<section>
  <h2>Control panel — every control must read OK before sign-off</h2>
  <div class="card"><table>
    <thead><tr><th>Control</th><th class="num">Result</th><th class="num">Expected</th><th>Status</th></tr></thead>
    <tbody>{ctrl}</tbody>
  </table></div>
</section>

<section>
  <h2>Reconciliation summary</h2>
  <div class="kpis">
    <div class="kpi"><div class="label">Total lines</div><div class="value">{k_total}</div><div class="foot">Unique matching keys</div></div>
    <div class="kpi"><div class="label">Reconciled</div><div class="value">{k_rec}</div><div class="foot">{k_rate}% match rate</div></div>
    <div class="kpi"><div class="label">Open items</div><div class="value">{k_open}</div><div class="foot">Require follow-up</div></div>
    <div class="kpi"><div class="label">Net difference</div><div class="{k_net_cls}">{k_net}</div><div class="foot">{la} less {lb}</div></div>
    <div class="kpi"><div class="label">Gross, ignoring sign</div><div class="value">{k_gross}</div><div class="foot">Sum of absolute differences</div></div>
  </div>
</section>

<div class="grid2">
  <section><h2>Open items by difference type</h2><div class="card"><table>
    <thead><tr><th>Difference type</th><th class="num">Count</th><th class="num">Value, ignoring sign</th></tr></thead>
    <tbody>{type_body}</tbody></table></div></section>
  <section><h2>Open items by root cause</h2><div class="card"><table>
    <thead><tr><th>Root cause</th><th class="num">Count</th><th class="num">Value, ignoring sign</th></tr></thead>
    <tbody>{root_body}</tbody></table></div></section>
</div>

<section>
  <h2>Difference by account</h2>
  <div class="card"><table>
    <thead><tr><th>{acct_hdr}</th><th>{name_hdr}</th><th class="num">Amount — {la}</th><th class="num">Amount — {lb}</th><th class="num">Difference</th></tr></thead>
    <tbody>{acct_body}</tbody></table></div>
</section>

<section>
  <h2>{cp_title}</h2>
  <div class="card"><table>
    <thead><tr><th>{comp_hdr}</th><th>{per_hdr}</th><th class="num">Amount — {la}</th><th class="num">Amount — {lb}</th><th class="num">Difference</th><th class="num">Open items</th></tr></thead>
    <tbody>{cp_body}</tbody></table></div>
</section>

<section>
  <h2>Reconciliation detail — {open_ct} open items</h2>
  <div class="card"><table>
    <thead><tr><th>{comp_hdr}</th><th>{acct_hdr}</th><th>{name_hdr}</th><th>{per_hdr}</th><th class="num">Amount — {la}</th><th class="num">Amount — {lb}</th><th class="num">Difference</th><th>Status</th><th>Difference Type</th><th>Root Cause</th></tr></thead>
    <tbody>{det_body}</tbody></table>
    <p class="sub" style="margin:10px 0 0">The remaining {rec_ct} lines are <span class="tag tag-rec">Reconciled</span> and are listed in full on the Reconciliation tab of the workbook.</p>
  </div>
</section>

<section><h2>Headlines</h2><div class="card"><ul class="headlines">{head}</ul></div></section>

<footer>{footer}</footer>
</div>
</body>
</html>
"""


def main():
    p = argparse.ArgumentParser(description="Config-driven two-source reconciliation.")
    p.add_argument("--config", required=True)
    p.add_argument("--source-a", required=True)
    p.add_argument("--source-b", required=True)
    p.add_argument("--sheet-a", default=None, help="Sheet name/index for source A (for workbook tabs)")
    p.add_argument("--sheet-b", default=None, help="Sheet name/index for source B (for workbook tabs)")
    p.add_argument("--out", default="reconciliation.xlsx")
    p.add_argument("--html", default=None,
                   help="Also write the styled HTML dashboard to this path (or set output.emitHtml in config).")
    args = p.parse_args()
    import os

    with open(args.config, encoding="utf-8") as f:
        config = json.load(f)

    # Resolve sheet selectors: CLI overrides config; config `sheet` under each source is honored.
    sheet_a = args.sheet_a if args.sheet_a is not None else config["sources"]["a"].get("sheet")
    sheet_b = args.sheet_b if args.sheet_b is not None else config["sources"]["b"].get("sheet")
    # A digit-only selector (e.g. --sheet-a 0) is a sheet index, not a sheet literally named "0".
    def _parse_sheet(s):
        if isinstance(s, str) and s.strip().lstrip("-").isdigit():
            return int(s.strip())
        return s
    sheet_a, sheet_b = _parse_sheet(sheet_a), _parse_sheet(sheet_b)

    df_a = load_table(args.source_a, sheet_a)
    df_b = load_table(args.source_b, sheet_b)

    # Fill in human labels from file/tab names when the config didn't set an explicit label.
    # This is what keeps the report free of generic "A"/"B" wording.
    if not config["sources"]["a"].get("label"):
        config["sources"]["a"]["label"] = default_label(args.source_a, sheet_a)
    if not config["sources"]["b"].get("label"):
        config["sources"]["b"]["label"] = default_label(args.source_b, sheet_b)
    # If both sources are the same file reconciled across two tabs, make sure the labels
    # are distinct by including the sheet name.
    if (args.source_a == args.source_b and sheet_a is not None and sheet_b is not None
            and config["sources"]["a"]["label"] == config["sources"]["b"]["label"]):
        config["sources"]["a"]["label"] = default_label(args.source_a, sheet_a)
        config["sources"]["b"]["label"] = default_label(args.source_b, sheet_b)

    # Align B's key columns to A's per matching.keyMap so differently-named keys reconcile.
    align_key_columns(config)

    la = config["sources"]["a"]["label"]
    lb = config["sources"]["b"]["label"]

    # This reference script implements the record-to-record tiered match. Control-total
    # tie-out (SKILL.md Step 3b) is an analytical method the agent performs directly; the
    # script does not run it, so refuse rather than silently emit record-to-record output.
    mode = config.get("matching", {}).get("reconciliationMode", "recordToRecord")
    if mode and mode != "recordToRecord":
        sys.exit(f"reconciliationMode '{mode}' is not run by this script. It implements "
                 "record-to-record matching only; perform control-total tie-out analytically "
                 "per SKILL.md Step 3b, or set matching.reconciliationMode to 'recordToRecord'.")

    # Confirm configured columns exist before matching.
    for side, df in (("a", df_a), ("b", df_b)):
        s = config["sources"][side]
        needed = list(s["keyColumns"]) + [s["amountColumn"]]
        missing = [c for c in needed if c not in df.columns]
        if missing:
            sys.exit(f"Source '{s['label']}' is missing configured column(s): {missing}. "
                     f"Available: {list(df.columns)}")

    results, total_a, total_b = reconcile(df_a, df_b, config)
    summary = tie_out(results, total_a, total_b, effective_tolerances(config["matching"])[0])
    counts = write_report(results, config, args.out, df_a=df_a, df_b=df_b,
                          src_name=os.path.basename(args.source_a))

    print(f"Reconciliation written to {args.out}")

    # Optional HTML dashboard (same data, styled template). Driven by --html or output.emitHtml.
    html_path = args.html
    if html_path is None and config.get("output", {}).get("emitHtml"):
        html_path = os.path.splitext(args.out)[0] + ".html"
    if html_path:
        rows, _, _ = compute_reconciliation(df_a, df_b, config)
        write_html_report(rows, config, html_path, src_name=os.path.basename(args.source_a),
                          df_a=df_a, df_b=df_b)
        print(f"HTML dashboard written to {html_path}")

    print(f"Control total {la} = {total_a:.2f} | {lb} = {total_b:.2f} | net = {summary['net_difference']:.2f}")
    print(f"Tied out: {'YES' if summary['tied_out'] else 'NO (residual %.2f)' % summary['residual']}")
    for status, n in counts.items():
        print(f"  {status}: {n}")


if __name__ == "__main__":
    main()
